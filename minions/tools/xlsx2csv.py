#coding: utf-8

"""
xlsx文件格式转csv格式
依赖openpyxl库：http://openpyxl.readthedocs.org/en/latest/

Author: hatlonely@gmail.com
Date: 2015-07-21
"""

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import csv
import os
import sys

def xlsx2csv(filename):
    try:
        xlsx_file_reader = load_workbook(filename=filename)
        for sheet in xlsx_file_reader.get_sheet_names():
            # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
            csv_filename = '{xlsx}_{sheet}.csv'.format(
                xlsx=os.path.splitext(filename.replace(' ', '_'))[0],
                sheet=sheet.replace(' ', '_'))

            csv_file = file(csv_filename, 'wb')
            csv_file_writer = csv.writer(csv_file)

            sheet_ranges = xlsx_file_reader[sheet]
            for row in sheet_ranges.rows:
                row_container = []
                for cell in row:
                    if type(cell.value) == unicode:
                        row_container.append(cell.value.encode('utf-8'))
                    else:
                        row_container.append(str(cell.value))
                csv_file_writer.writerow(row_container)
            csv_file.close()

    except Exception as e:
        print(e)

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('usage: xlsx2csv <xlsx file name>')
    else:
        xlsx2csv(sys.argv[1])
    sys.exit(0)
